import tkinter as tk
from tkinter import filedialog

from pathlib import Path
from openpyxl import load_workbook, Workbook


# variables defined

number_of_files = 'No files Selected'  # to display the number of excel sheets
number_of_cells = 'No cells Selected'  # to display the number of cells selected
fileloc = ''
savelabel = ''
folderpath = ''
file_loc_list = []  # to save the files selected
header_list = []  # to save the headers
cell_list = []  # to save the cells entered


def openFolder():
    header_list.clear()
    cell_list.clear()
    file_loc_list.clear()
    heading_button.config(state=tk.DISABLED)
    cell_clear_button.config(state=tk.DISABLED)
    cell_button.config(state=tk.DISABLED)
    selected_cell_label_display.config(text="")
    ready_label.config(text="")
    extract_button.config(state=tk.DISABLED)
    final_label.config(text="")
    global folderpath
    folderpath = filedialog.askdirectory()
    if folderpath:
        count = Path(folderpath)
        warning_label1.config(text='Total ' + str(len(list(count.rglob("*xlsx*")))) + ' excel files added.',
                              fg="#B03A2E",
                              bg="#F4F6F6", )
        heading_button.config(state=tk.NORMAL)
        cell_button.config(state=tk.NORMAL)
        cell_clear_button.config(state=tk.NORMAL)
        clear_files_button.config(state=tk.NORMAL)

    else:
        warning_label1.config(text="No file selected", bg="yellow", )
        heading_button.config(state=tk.DISABLED)
        cell_button.config(state=tk.DISABLED)
        cell_clear_button.config(state=tk.DISABLED)
        clear_files_button.config(state=tk.DISABLED)


def openFile():
    file_loc_list.clear()
    header_list.clear()
    cell_list.clear()
    file_loc_list.clear()
    heading_button.config(state=tk.DISABLED)
    cell_clear_button.config(state=tk.DISABLED)
    cell_button.config(state=tk.DISABLED)
    selected_cell_label_display.config(text="No Cells Selected")
    ready_label.config(text="")
    extract_button.config(state=tk.DISABLED)
    final_label.config(text="")
    filepath = filedialog.askopenfilenames()
    for i in filepath:
        if i.endswith('.xlsx'):
            file_loc_list.append(i)

    if file_loc_list:
        count = len(file_loc_list)
        if count == 1:
            warning_label1.config(text="Total " + str(count) + " excel file added",
                                  fg="#B03A2E", bg="#F4F6F6", )
            heading_button.config(state=tk.NORMAL)
            cell_button.config(state=tk.NORMAL)
            cell_clear_button.config(state=tk.NORMAL)
            clear_files_button.config(state=tk.NORMAL)

        else:
            warning_label1.config(text="Total " + str(count) + " excel files added",
                                  fg="#B03A2E", bg="#F4F6F6", )
            heading_button.config(state=tk.NORMAL)
            cell_button.config(state=tk.NORMAL)
            cell_clear_button.config(state=tk.NORMAL)
            clear_files_button.config(state=tk.NORMAL)
    else:
        warning_label1.config(text="No file selected", bg="yellow", )
        heading_button.config(state=tk.DISABLED)
        cell_button.config(state=tk.DISABLED)
        cell_clear_button.config(state=tk.DISABLED)
        clear_files_button.config(state=tk.DISABLED)


# clearing the list of files already selected
def clearFiles():
    file_loc_list.clear()
    warning_label1.config(text="No file selected", bg="yellow", )
    heading_button.config(state=tk.DISABLED)
    cell_clear_button.config(state=tk.DISABLED)
    cell_button.config(state=tk.DISABLED)
    extract_button.config(state=tk.DISABLED)


# function to collect the headers

def getHeaders():
    header_list.clear()
    header = heading_entry.get().title().split(",")
    if header:
        header_list.extend(header)


# function to collect the cells to extract data

def getCells():
    cells = cell_entry.get().upper().split(",")
    selected_cell_label_display.config(text=cells)
    if cells:
        extract_button.config(state=tk.NORMAL)
        ready_label.config(text="Click button to extract")
        cell_list.extend(cells)


# function to clear the cells displayed
def clearCells():
    cell_entry.delete(0, "end")
    cell_list.clear()
    extract_button.config(state=tk.DISABLED)
    selected_cell_label_display.config(text=number_of_cells)


# function for header entry placeholder
def heading_on_entry_click(event):
    if heading_entry.get() == "E.g.: Name,Email,Place . . .":
        heading_entry.delete(0, tk.END)
        heading_entry.configure(fg="black")


def heading_on_focus_out(event):
    if heading_entry.get() == "":
        heading_entry.insert(0, "E.g.: Name,Email,Place . . .")
        heading_entry.configure(fg="#99A3A4")


# function for cell entry placeholder
def cell_on_entry_click(event):
    if cell_entry.get() == "E.g.: A1,A2,A3,B1,B2 . . .":
        cell_entry.delete(0, tk.END)
        cell_entry.configure(fg="black")


def cell_on_focus_out(event):
    if cell_entry.get() == "":
        cell_entry.insert(0, "E.g.: A1,A2,A3,B1,B2 . . .")
        cell_entry.configure(fg="#99A3A4")


def extractData():
    filelocation_to_save = filedialog.asksaveasfilename(defaultextension='.xlsx')
    extracting_list = []
    if header_list:
        extracting_list.append(tuple(header_list))
    # loading working books if files added
    if file_loc_list:
        for i in file_loc_list:
            wb = load_workbook(i)
            ws = wb.active
            # extracting datas if cell exists
            if cell_list:
                singlesheet = []
                for j in cell_list:
                    data = ws[j].value
                    singlesheet.append(data)
                extracting_list.append(tuple(singlesheet))
            else:
                extract_button.config(state=tk.DISABLED)
    elif folderpath:
        directory = Path(folderpath)
        files = list(directory.rglob("*.xlsx*"))
        # extracting datas
        for i in files:
            wb = load_workbook(i)
            ws = wb.active
            # extracting datas if cell exists
            if cell_list:
                singleshheet = []
                for j in cell_list:
                    data = ws[j].value
                    singleshheet.append(data)
                extracting_list.append(tuple(singleshheet))
            else:
                extract_button.config(state=tk.DISABLED)
    else:
        extract_button.config(state=tk.DISABLED)

    if extracting_list:
        wb_new = Workbook()
        ws_new = wb_new.active
        ws_new.title = "Extracted Sheet"

        # passing values to new created workbook
        for i in extracting_list:
            ws_new.append(i)
        wb_new.save(filelocation_to_save)
        final_label.config(text="Successfully new file created")
    else:
        final_label.config(text="Data extraction filed")


"""GUI Start here"""
window = tk.Tk()
window.title("Extract Datas")
window.geometry("1500x1050")
window['bg'] = "#34495E"

label = tk.Label(window, text="EASY YOUR EXCEL", bg="#85929E",
                 font=("Comic Sans MS", 17, "bold",), height=2, padx=70)
label.place(x=50)

# frame1
frame_1 = tk.Frame(window, bg="#AEB6BF", width=362, height=700)
frame_1.pack_propagate(False)
frame_1.place(x=50, y=100)

# file browsing lobel and button
file_label = tk.Label(frame_1, text="Select your Excel file", font=("Times", 12),
                      bg="#AEB6BF",
                      fg="#1B2631")
file_label.place(x=30, y=40)

file_button = tk.Button(frame_1, text="Browse File", font=("Georgia", 12),
                        bd=1, relief="raised",
                        bg="#283747",
                        fg="#AEB6BF",
                        padx=5,
                        command=lambda: openFile())
file_button.place(x=230, y=36)

# folder browsing label and button
folder_label = tk.Label(frame_1, text="Multiple files? Select folder", font=("Times", 12),
                        fg="#1B2631",
                        bg="#AEB6BF",
                        )
folder_label.place(x=30, y=100)

folder_button = tk.Button(frame_1, text="Browse Folder", font=("Georgia", 12,),
                          bd=1, relief="raised",
                          bg="#283747",
                          fg="#AEB6BF",
                          padx=5,
                          command=lambda: openFolder(),
                          )
folder_button.place(x=210, y=96)

warning_label1 = tk.Label(frame_1, text=number_of_files,
                          font=("Times", 13),
                          bg="#F4F6F6",
                          fg="#85929E",
                          width=32, pady=2)
warning_label1.place(x=30, y=165)

# clearing the selected files
clear_files_button = tk.Button(frame_1, text="clear sheets", font=("Georgia", 12,),
                               bd=1, relief="solid",
                               padx=5,
                               bg="#D6DBDF",
                               fg="#283747",
                               state=tk.DISABLED,
                               command=lambda: clearFiles(),
                               )
clear_files_button.place(x=220, y=220)

# frame2
frame_2 = tk.Frame(window, bg="#AEB6BF", width=700, height=550)
frame_2.pack_propagate(False)
frame_2.place(x=500, y=70)

# heading part

head_label = tk.Label(frame_2, text="Make your heading",
                      font=("Times", 14),
                      bg="#AEB6BF",
                      )
head_label.place(x=30, y=50)

heading_entry = tk.Entry(frame_2,
                         width=50,
                         fg="#99A3A4",
                         font=("Times", 13))

heading_entry.insert(0, "E.g.: Name,Email,Place . . .")
heading_entry.bind("<FocusIn>", heading_on_entry_click)
heading_entry.bind("<FocusOut>", heading_on_focus_out)
heading_entry.place(x=200, y=50)

heading_button = tk.Button(frame_2, text="Add Headers", font=("Georgia", 12,),
                           bd=1, relief="raised",
                           bg="#283747",
                           fg="#AEB6BF",
                           padx=10,
                           state=tk.DISABLED,
                           command=lambda: getHeaders(),
                           )

heading_button.place(x=528, y=110)

# selecting cells

cell_label = tk.Label(frame_2, text="Required cells",
                      font=("Times", 14),
                      pady=1,
                      bg="#AEB6BF")
cell_label.place(x=30, y=200)

cell_entry = tk.Entry(frame_2,
                      width=50,
                      fg="#99A3A4",
                      font=("Times", 13))

cell_entry.insert(0, "E.g.: A1,A2,A3,B1,B2 . . .")
cell_entry.bind("<FocusIn>", cell_on_entry_click)
cell_entry.bind("<FocusOut>", cell_on_focus_out)
cell_entry.place(x=200, y=200)

cell_clear_button = tk.Button(frame_2,
                              text="Clear",
                              font=("Georgia", 12),
                              bd=1, relief="solid",
                              padx=10,
                              bg="#D6DBDF",
                              fg="#283747",
                              state=tk.DISABLED,
                              command=lambda: clearCells(),

                              )

cell_clear_button.place(x=500, y=260)

cell_button = tk.Button(frame_2, text="Done", font=("Georgia", 12,),
                        bd=1, relief="raised",
                        bg="#283747",
                        fg="#AEB6BF",
                        padx=10,
                        state=tk.DISABLED,
                        command=lambda: getCells(),
                        )

cell_button.place(x=585, y=260)

selected_cell_label = tk.Label(frame_2, text="Selected cells :",
                               font=("Times", 14),
                               pady=1,
                               bg="#AEB6BF")

selected_cell_label.place(x=30, y=340)

selected_cell_label_display = tk.Label(frame_2, text=number_of_cells,
                                       font=("Times", 13),
                                       bg="#EAEDED",
                                       fg="#85929E",
                                       width=49, pady=2)
selected_cell_label_display.place(x=200, y=340)

# ready label and extract button

ready_label = tk.Label(frame_2, text="",
                       font=("Times", 14),
                       pady=1,
                       bg="#AEB6BF")
ready_label.place(x=320, y=420)

extract_button = tk.Button(frame_2,
                           text="Extract",
                           font=("Georgia", 14),
                           bd=1, relief="ridge",
                           padx=15, pady=1,
                           bg="#212F3C",
                           fg="#AEB6BF",
                           state=tk.DISABLED,
                           command=lambda: extractData(),
                           )

extract_button.place(x=545, y=410)

# Final label

final_label = tk.Label(frame_2, text='',
                       font=("Times", 14),
                       bg="#AEB6BF"
                       )
final_label.place(x=400, y=480)

window.mainloop()

"""GUI Ends here"""
